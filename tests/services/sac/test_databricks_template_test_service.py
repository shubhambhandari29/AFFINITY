import asyncio
from io import BytesIO

import pytest
from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from services.loss_run import databricks_template_test_service


def _template_bytes() -> bytes:
    workbook = Workbook()
    workbook.active.title = "Cover Page"

    claims = workbook.create_sheet("Claims Data")
    claims.append(["Claim Number"])
    claims.append(["123"])
    claims_table = Table(displayName="ClaimsData", ref="A1:A2")
    claims_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
    claims.add_table(claims_table)

    record_only = workbook.create_sheet("Record Only")
    record_only.append(["Claim Number"])
    record_only.append(["456"])
    record_only_table = Table(displayName="RecordOnlyData", ref="A1:A2")
    record_only_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
    record_only.add_table(record_only_table)

    workbook.create_sheet("Summary By Policy Year")
    workbook.create_sheet("Chart")

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_local_access_downloads_validates_and_uploads(monkeypatch):
    template_bytes = _template_bytes()
    captured = {}

    class Download:
        contents = BytesIO(template_bytes)

    class Files:
        @staticmethod
        def download(path):
            assert path == databricks_template_test_service.DATABRICKS_TEMPLATE_PATH
            return Download()

        @staticmethod
        def upload(path, contents, *, overwrite):
            captured["path"] = path
            captured["bytes"] = contents.read()
            assert overwrite is False

    class Client:
        files = Files()

    class WorkspaceClient:
        def __new__(cls, *, host, profile):
            assert host == databricks_template_test_service.DATABRICKS_HOST
            assert profile == databricks_template_test_service.DATABRICKS_LOCAL_PROFILE
            return Client()

    monkeypatch.setitem(
        __import__("sys").modules,
        "databricks.sdk",
        type("SDK", (), {"WorkspaceClient": WorkspaceClient}),
    )

    result = databricks_template_test_service._test_local_access()

    assert result["status"] == "success"
    assert result["authenticationMode"] == "local-oauth-profile"
    assert result["testOutputPath"] == captured["path"]
    assert captured["bytes"] == template_bytes


def test_azure_access_downloads_validates_and_uploads(monkeypatch):
    template_bytes = _template_bytes()
    requests_made = []

    class Credential:
        @staticmethod
        def get_token(scope):
            assert scope == databricks_template_test_service.DATABRICKS_TOKEN_SCOPE
            return type("AccessToken", (), {"token": "test-token"})()

    class Response:
        def __init__(self, content=b"", status_code=200):
            self.content = content
            self.status_code = status_code
            self.ok = status_code < 400
            self.text = ""

    def fake_get(url, *, headers, timeout):
        requests_made.append(("GET", url))
        assert headers == {"Authorization": "Bearer test-token"}
        assert timeout == 60
        return Response(template_bytes)

    def fake_put(url, *, headers, params, data, timeout):
        requests_made.append(("PUT", url))
        assert headers == {
            "Authorization": "Bearer test-token",
            "Content-Type": "application/octet-stream",
        }
        assert params == {"overwrite": "false"}
        assert data == template_bytes
        assert timeout == 60
        return Response(status_code=204)

    monkeypatch.setattr(
        databricks_template_test_service,
        "ManagedIdentityCredential",
        Credential,
    )
    monkeypatch.setattr(databricks_template_test_service.requests, "get", fake_get)
    monkeypatch.setattr(databricks_template_test_service.requests, "put", fake_put)

    result = databricks_template_test_service._test_azure_access()

    assert result["status"] == "success"
    assert result["authenticationMode"] == "system-assigned-managed-identity"
    assert result["testOutputPath"].startswith(
        databricks_template_test_service.DATABRICKS_OUTPUT_DIR
    )
    assert [method for method, _ in requests_made] == ["GET", "PUT"]


def test_databricks_error_includes_safe_response_detail():
    class Response:
        ok = False
        status_code = 403
        text = '{"error_code":"PERMISSION_DENIED"}'

        @staticmethod
        def json():
            return {
                "error_code": "PERMISSION_DENIED",
                "message": "Missing READ VOLUME",
            }

    with pytest.raises(RuntimeError) as error:
        databricks_template_test_service._raise_databricks_error(
            Response(),
            "template read",
        )

    assert str(error.value) == (
        "Databricks template read failed with HTTP 403: Missing READ VOLUME"
    )


@pytest.mark.parametrize(
    ("service_name", "worker_name", "expected_error"),
    [
        (
            "test_local_databricks_access",
            "_test_local_access",
            "Local Databricks read/write test failed",
        ),
        (
            "test_azure_databricks_access",
            "_test_azure_access",
            "Azure managed-identity Databricks read/write test failed",
        ),
    ],
)
def test_service_returns_safe_http_error(
    monkeypatch,
    service_name,
    worker_name,
    expected_error,
):
    def fail():
        raise RuntimeError("test failure")

    monkeypatch.setattr(databricks_template_test_service, worker_name, fail)

    with pytest.raises(HTTPException) as error:
        asyncio.run(getattr(databricks_template_test_service, service_name)())

    assert error.value.status_code == 500
    assert error.value.detail["error"] == expected_error
