import asyncio
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image as PILImage

from services.loss_run import loss_run_service


def _make_template(path):
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover Page"
    logo_path = path.parent / "test-logo.png"
    PILImage.new("RGB", (4, 4), "orange").save(logo_path)
    cover.add_image(ExcelImage(logo_path), "A1")

    for sheet_name, table_name in (
        ("Claims Data", "ClaimsData"),
        ("Record Only", "RecordOnlyData"),
    ):
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(["Old Column", "Old Value"])
        worksheet.append(["old", "old"])
        excel_table = Table(displayName=table_name, ref="A1:B2")
        excel_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        worksheet.add_table(excel_table)

    workbook.create_sheet("Summary By Policy Year")
    workbook.create_sheet("Chart")
    workbook.save(path)


def test_create_workbook_populates_claims_record_only_and_cover(tmp_path):
    template_path = tmp_path / "SACLossRunTemplate.xlsx"
    _make_template(template_path)

    records = [
        {
            "Customer Number": "00123",
            "Claim Number": "C-1",
            "Exposure": 1,
            "Record Only Indicator": "N",
            "Total Incurred": 2500,
        },
        {
            "Customer Number": "00123",
            "Claim Number": "C-2",
            "Exposure": 2,
            "Record Only Indicator": "Y",
            "Total Incurred": 0,
        },
    ]

    workbook_bytes = loss_run_service._create_workbook(
        records,
        "00123",
        "Example Customer",
        template_path.read_bytes(),
    )

    workbook = load_workbook(BytesIO(workbook_bytes))
    assert workbook["Cover Page"]["B2"].value == "00123"
    assert workbook["Cover Page"]["B3"].value == "Example Customer"
    assert len(workbook["Cover Page"]._images) == 1
    assert workbook["Claims Data"]["B2"].value == "C-1"
    assert workbook["Claims Data"]["C2"].value == "01"
    assert workbook["Record Only"]["B2"].value == "C-2"
    assert workbook["Record Only"]["C2"].value == "02"
    assert "Record Only Indicator" not in [
        cell.value for cell in workbook["Claims Data"][1]
    ]
    assert "Total Incurred" not in [cell.value for cell in workbook["Record Only"][1]]
    workbook.close()


def test_generate_selected_loss_runs_handles_one_or_more_customers(monkeypatch):
    template_bytes = b"template"
    uploaded = []

    class Storage:
        @staticmethod
        def download_template():
            return template_bytes

        @staticmethod
        def upload_report(filename, workbook_bytes):
            uploaded.append((filename, workbook_bytes))

    monkeypatch.setattr(loss_run_service, "DatabricksLossRunStorage", Storage)

    calls = []

    async def fake_query(query, params):
        calls.append((query, params))
        if "tblAcctSpecial" in query:
            return [
                {"CustomerNum": "00123", "CustomerName": "Example/Customer"},
                {"CustomerNum": "00456", "CustomerName": "No Data Customer"},
            ]
        return [
            {
                "Customer Number": "00123",
                "Claim Number": "C-1",
                "Record Only Indicator": "N",
            }
        ]

    def fake_create(records, customer_num, customer_name, received_template):
        assert received_template == template_bytes
        return b"workbook"

    async def fake_threadpool(func, *args):
        return func(*args)

    monkeypatch.setattr(loss_run_service, "run_raw_query_async", fake_query)
    monkeypatch.setattr(loss_run_service, "_create_workbook", fake_create)
    monkeypatch.setattr(loss_run_service, "run_in_threadpool", fake_threadpool)

    result = asyncio.run(
        loss_run_service.generate_loss_runs(["00123", "00456", "00123"])
    )

    assert result["requestedCount"] == 2
    assert result["generatedCount"] == 1
    assert result["failedCount"] == 1
    assert result["failures"] == [
        {"customerNumber": "00456", "reason": "No loss-run records found"}
    ]
    assert "files" not in result
    assert len(calls) == 2
    assert calls[0][1] == ["00123", "00456"]
    assert calls[1][1] == ["00123", "00456"]
    assert "WHERE [Customer Number] IN (?, ?)" in calls[1][0]
    assert uploaded[0][0].startswith("Example_Customer_")
    assert uploaded[0][1] == b"workbook"


def test_generate_all_loss_runs_uses_current_eligibility_rules(monkeypatch):
    class Storage:
        @staticmethod
        def download_template():
            return b"template"

        @staticmethod
        def upload_report(filename, workbook_bytes):
            return None

    monkeypatch.setattr(loss_run_service, "DatabricksLossRunStorage", Storage)

    calls = []

    async def fake_query(query, params=None):
        calls.append((query, params))
        if "tblAcctSpecial" in query:
            return [{"CustomerNum": "00123", "CustomerName": "Example Customer"}]
        return [
            {
                "Customer Number": "00123",
                "Claim Number": "C-1",
                "Record Only Indicator": "N",
            }
        ]

    def fake_create(records, customer_num, customer_name, template_bytes):
        return b"workbook"

    async def fake_threadpool(func, *args):
        return func(*args)

    monkeypatch.setattr(loss_run_service, "run_raw_query_async", fake_query)
    monkeypatch.setattr(loss_run_service, "_create_workbook", fake_create)
    monkeypatch.setattr(loss_run_service, "run_in_threadpool", fake_threadpool)

    result = asyncio.run(loss_run_service.generate_loss_runs())

    assert result["generatedCount"] == 1
    assert "AcctStatus = 'Active'" in calls[0][0]
    assert "LossRunDistFreq <> 'Not Needed'" in calls[0][0]
    assert calls[1][0].strip() == "SELECT * FROM dbo.SAC_Loss_Run"


def test_generate_loss_runs_records_upload_failure(monkeypatch):
    class Storage:
        @staticmethod
        def download_template():
            return b"template"

        @staticmethod
        def upload_report(filename, workbook_bytes):
            raise RuntimeError("upload failed")

    async def fake_query(query, params):
        if "tblAcctSpecial" in query:
            return [{"CustomerNum": "00123", "CustomerName": "Example Customer"}]
        return [
            {
                "Customer Number": "00123",
                "Claim Number": "C-1",
                "Record Only Indicator": "N",
            }
        ]

    def fake_create(records, customer_num, customer_name, template_bytes):
        return b"workbook"

    async def fake_threadpool(func, *args):
        return func(*args)

    monkeypatch.setattr(loss_run_service, "DatabricksLossRunStorage", Storage)
    monkeypatch.setattr(loss_run_service, "run_raw_query_async", fake_query)
    monkeypatch.setattr(loss_run_service, "_create_workbook", fake_create)
    monkeypatch.setattr(loss_run_service, "run_in_threadpool", fake_threadpool)

    result = asyncio.run(loss_run_service.generate_loss_runs(["00123"]))

    assert result == {
        "requestedCount": 1,
        "generatedCount": 0,
        "failedCount": 1,
        "failures": [
            {
                "customerNumber": "00123",
                "reason": "Failed to generate or upload workbook",
            }
        ],
    }
