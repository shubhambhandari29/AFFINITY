import logging
import re
from datetime import datetime
from io import BytesIO

import pandas as pd
from fastapi import HTTPException
from fastapi.concurrency import run_in_threadpool
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import TableColumn

from core.db_helpers import run_raw_query_async
from services.loss_run.databricks_storage_service import DatabricksLossRunStorage

logger = logging.getLogger(__name__)

RECORD_ONLY_EXCLUDED_COLUMNS = [
    "Claims above 50K",
    "Total Incurred",
    "Total Incurred + ALAE",
    "Total Paid Loss Net Salvage/Subro/Loss Recovery",
    "Incurred w/o ALAE",
    "ALAE Reserve",
    "ALAE Paid",
    "Salvage Recovery",
    "Subro Recovery",
    "Loss Recovery",
    "Deductible Recovery",
    "Expense Recovery",
    "Litigation Status",
    "Outstanding Loss Reserve",
]


def _write_excel_table(worksheet, table_name: str, dataframe: pd.DataFrame) -> None:
    excel_table = worksheet.tables[table_name]
    old_cells = worksheet[excel_table.ref]
    row_count = len(dataframe) + 1
    column_count = len(dataframe.columns)
    new_ref = f"A1:{get_column_letter(column_count)}{row_count}"

    for row_index, row in enumerate(worksheet[new_ref]):
        for column_index, cell in enumerate(row):
            if row_index == 0:
                cell.value = dataframe.columns[column_index]
                continue

            value = dataframe.iat[row_index - 1, column_index]
            cell.value = None if pd.isna(value) else value

    for row in old_cells:
        for cell in row:
            if cell.row > row_count or cell.column > column_count:
                cell.value = None

    excel_table.tableColumns = [
        TableColumn(id=index + 1, name=column)
        for index, column in enumerate(dataframe.columns)
    ]
    excel_table.autoFilter = None
    excel_table.ref = new_ref


def _create_workbook(
    records: list[dict], customer_num: str, customer_name: str, template_bytes: bytes
) -> bytes:
    dataframe = pd.DataFrame(records)

    if "Exposure" in dataframe.columns:
        dataframe["Exposure"] = dataframe["Exposure"].apply(
            lambda value: f"{int(value):02d}"
            if pd.notna(value) and str(value).strip()
            else ""
        )

    dataframe["Distinct Claim Helper"] = (
        ~dataframe["Claim Number"].duplicated()
    ).astype(int)

    record_only = dataframe[dataframe["Record Only Indicator"] == "Y"].drop(
        columns=["Record Only Indicator", *RECORD_ONLY_EXCLUDED_COLUMNS],
        errors="ignore",
    )
    claims = dataframe[dataframe["Record Only Indicator"] != "Y"].drop(
        columns=["Record Only Indicator"]
    )

    if record_only.empty:
        record_only = pd.DataFrame(
            [["No Records Only", *([None] * (len(record_only.columns) - 1))]],
            columns=record_only.columns,
        )

    workbook = load_workbook(BytesIO(template_bytes))
    _write_excel_table(
        workbook["Claims Data"], "ClaimsData", claims.reset_index(drop=True)
    )
    _write_excel_table(
        workbook["Record Only"], "RecordOnlyData", record_only.reset_index(drop=True)
    )

    cover_page = workbook["Cover Page"]
    cover_page.cell(2, 2, customer_num)
    cover_page.cell(3, 2, customer_name)
    cover_page.cell(4, 2, datetime.now().strftime("%m/%d/%Y"))

    for sheet_name in ("Summary By Policy Year", "Chart"):
        if sheet_name not in workbook.sheetnames:
            continue
        for pivot in getattr(workbook[sheet_name], "_pivots", []):
            pivot.cache.refreshOnLoad = True

    workbook.calculation.fullCalcOnLoad = True
    output = BytesIO()
    try:
        workbook.save(output)
        return output.getvalue()
    finally:
        workbook.close()


async def generate_loss_runs(customer_nums: list[str] | None = None) -> dict:
    try:
        storage = DatabricksLossRunStorage()
        template_bytes = await run_in_threadpool(storage.download_template)

        if customer_nums is None:
            customers = await run_raw_query_async(
                """
                SELECT CustomerNum, CustomerName
                FROM dbo.tblAcctSpecial
                WHERE AcctStatus = 'Active'
                  AND LossRunDistFreq <> 'Not Needed'
                  AND LossRunDistFreq <> ''
                """
            )
            requested_numbers = list(
                dict.fromkeys(
                    str(customer["CustomerNum"]).strip() for customer in customers
                )
            )

            records = await run_raw_query_async("SELECT * FROM dbo.SAC_Loss_Run")
        else:
            requested_numbers = list(
                dict.fromkeys(
                    str(customer_num).strip()
                    for customer_num in customer_nums
                    if str(customer_num).strip()
                )
            )
            if not requested_numbers:
                raise HTTPException(
                    status_code=400,
                    detail={"error": "At least one customer number is required"},
                )

            placeholders = ", ".join("?" for _ in requested_numbers)
            customers = await run_raw_query_async(
                f"""
                SELECT CustomerNum, CustomerName
                FROM dbo.tblAcctSpecial
                WHERE CustomerNum IN ({placeholders})
                """,
                requested_numbers,
            )

            records = await run_raw_query_async(
                f"""
                SELECT *
                FROM dbo.SAC_Loss_Run
                WHERE [Customer Number] IN ({placeholders})
                """,
                requested_numbers,
            )

        customer_names = {
            str(customer["CustomerNum"]).strip(): str(
                customer.get("CustomerName") or customer["CustomerNum"]
            ).strip()
            for customer in customers
        }
        records_by_customer: dict[str, list[dict]] = {}
        for record in records:
            customer_num = str(record.get("Customer Number") or "").strip()
            records_by_customer.setdefault(customer_num, []).append(record)

        generated_count = 0
        failures = []

        for customer_num in requested_numbers:
            customer_name = customer_names.get(customer_num)
            customer_records = records_by_customer.get(customer_num)

            if customer_name is None:
                failures.append(
                    {"customerNumber": customer_num, "reason": "Customer not found"}
                )
                continue
            if not customer_records:
                failures.append(
                    {
                        "customerNumber": customer_num,
                        "reason": "No loss-run records found",
                    }
                )
                continue

            safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", customer_name).strip(" .")
            filename = f"{safe_name or customer_num}_{datetime.now():%Y_%m_%d}.xlsx"

            try:
                workbook_bytes = await run_in_threadpool(
                    _create_workbook,
                    customer_records,
                    customer_num,
                    customer_name,
                    template_bytes,
                )
                await run_in_threadpool(storage.upload_report, filename, workbook_bytes)
                generated_count += 1
            except Exception:
                logger.exception(
                    "Failed to generate or upload loss-run workbook for customer %s",
                    customer_num,
                )
                failures.append(
                    {
                        "customerNumber": customer_num,
                        "reason": "Failed to generate or upload workbook",
                    }
                )

        return {
            "requestedCount": len(requested_numbers),
            "generatedCount": generated_count,
            "failedCount": len(failures),
            "failures": failures,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to generate loss-run reports")
        raise HTTPException(
            status_code=500, detail={"error": "Failed to generate loss-run reports"}
        ) from exc
