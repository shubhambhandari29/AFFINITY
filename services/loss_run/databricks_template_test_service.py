from datetime import UTC, datetime
from io import BytesIO
from urllib.parse import quote
from uuid import uuid4

import requests
from azure.identity import ManagedIdentityCredential
from fastapi import HTTPException
from fastapi.concurrency import run_in_threadpool
from openpyxl import load_workbook

DATABRICKS_HOST = "https://adb-8608532567795739.19.azuredatabricks.net"
DATABRICKS_LOCAL_PROFILE = "claims-preprod"
DATABRICKS_TOKEN_SCOPE = "2ff814a6-3304-4ab8-85cb-cd0e6f879c1d/.default"
DATABRICKS_TEMPLATE_PATH = "/Volumes/claims_data_pre_prod/gold/statics/SACLossRunTemplate.xlsx"
DATABRICKS_OUTPUT_DIR = (
    "/Volumes/claims_data_pre_prod/gold/external_volume/" "specialaccounts_lossruns_temporary"
)

REQUIRED_SHEETS = {
    "Cover Page",
    "Claims Data",
    "Record Only",
    "Summary By Policy Year",
    "Chart",
}


def _validate_template(template_bytes: bytes) -> dict:
    workbook = load_workbook(BytesIO(template_bytes))
    try:
        missing_sheets = sorted(REQUIRED_SHEETS.difference(workbook.sheetnames))
        claims_table_found = "ClaimsData" in workbook["Claims Data"].tables
        record_only_table_found = "RecordOnlyData" in workbook["Record Only"].tables
        cover_image_count = len(workbook["Cover Page"]._images)

        if missing_sheets or not claims_table_found or not record_only_table_found:
            raise ValueError("The downloaded file is not the expected loss-run template")

        return {
            "fileSizeBytes": len(template_bytes),
            "sheetNames": workbook.sheetnames,
            "claimsDataTableFound": claims_table_found,
            "recordOnlyDataTableFound": record_only_table_found,
            "coverImageCount": cover_image_count,
        }
    finally:
        workbook.close()


def _test_output_path(authentication_mode: str) -> str:
    timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    suffix = uuid4().hex[:8]
    filename = f"SACPlatformAPI_{authentication_mode}_access_test_{timestamp}_{suffix}.xlsx"
    return f"{DATABRICKS_OUTPUT_DIR}/{filename}"


def _raise_databricks_error(response: requests.Response, operation: str) -> None:
    if response.ok:
        return

    try:
        error = response.json()
        detail = error.get("message") or error.get("error_code") or response.text
    except ValueError:
        detail = response.text

    raise RuntimeError(f"Databricks {operation} failed with HTTP {response.status_code}: {detail}")


def _test_local_access() -> dict:
    from databricks.sdk import WorkspaceClient

    client = WorkspaceClient(
        host=DATABRICKS_HOST,
        profile=DATABRICKS_LOCAL_PROFILE,
    )

    download = client.files.download(DATABRICKS_TEMPLATE_PATH)
    with download.contents as stream:
        template_bytes = stream.read()

    output_path = _test_output_path("local-oauth")
    client.files.upload(
        output_path,
        BytesIO(template_bytes),
        overwrite=False,
    )

    return {
        "status": "success",
        "message": "Databricks template read and test-file write succeeded",
        "authenticationMode": "local-oauth-profile",
        "templatePath": DATABRICKS_TEMPLATE_PATH,
        "testOutputPath": output_path,
        **_validate_template(template_bytes),
    }


def _test_azure_access() -> dict:
    access_token = ManagedIdentityCredential().get_token(DATABRICKS_TOKEN_SCOPE)
    headers = {"Authorization": f"Bearer {access_token.token}"}

    template_url = (
        f"{DATABRICKS_HOST}/api/2.0/fs/files" f"{quote(DATABRICKS_TEMPLATE_PATH, safe='/')}"
    )
    download = requests.get(template_url, headers=headers, timeout=60)
    _raise_databricks_error(download, "template read")
    template_bytes = download.content

    output_path = _test_output_path("managed-identity")
    output_url = f"{DATABRICKS_HOST}/api/2.0/fs/files" f"{quote(output_path, safe='/')}"
    upload = requests.put(
        output_url,
        headers={
            **headers,
            "Content-Type": "application/octet-stream",
        },
        params={"overwrite": "false"},
        data=template_bytes,
        timeout=60,
    )
    _raise_databricks_error(upload, "test-file write")

    return {
        "status": "success",
        "message": "Databricks template read and test-file write succeeded",
        "authenticationMode": "system-assigned-managed-identity",
        "templatePath": DATABRICKS_TEMPLATE_PATH,
        "testOutputPath": output_path,
        **_validate_template(template_bytes),
    }


async def test_local_databricks_access() -> dict:
    try:
        return await run_in_threadpool(_test_local_access)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "error": "Local Databricks read/write test failed",
                "reason": str(exc),
            },
        ) from exc


async def test_azure_databricks_access() -> dict:
    try:
        return await run_in_threadpool(_test_azure_access)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail={
                "error": "Azure managed-identity Databricks read/write test failed",
                "reason": str(exc),
            },
        ) from exc
