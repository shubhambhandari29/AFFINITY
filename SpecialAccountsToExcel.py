# Databricks notebook source
# DBTITLE 1,Install openpyxl
# MAGIC %pip install --upgrade openpyxl
# MAGIC %restart_python

# COMMAND ----------

# Check access to Unity Catalog volume
try:
    files = dbutils.fs.ls('/Volumes/claims_data_pre_prod/gold/statics/')
    print(f"UC volume access OK: {len(files)} files found")
except Exception as e:
    raise RuntimeError(f"UC volume access failed: {e}")

# COMMAND ----------

# DBTITLE 1,Setup
# sql imports
from pyspark.sql import functions as F
from pyspark.sql.functions import col, regexp_replace
# package imports
import pandas as pd
import numpy as np
import os
from datetime import datetime
from shutil import copyfile
from openpyxl import Workbook
from openpyxl.worksheet import table
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# pyarrow plz
# spark.conf.set("spark.sql.execution.arrow.pyspark.enabled", "true")

# COMMAND ----------

# DBTITLE 1,Function Definition for Loss Run Generation
def createLossRun(df, data):
    custNum, custName = data
# , distFreq, recipCat, distVia, attnTo, email
    print('CREATING LOSS RUN EXCEL DOCUMENT FOR CUSTOMER:', custNum, custName)

    # Split between claim_df and record_only_df
    ro_df = df[df['Record Only Indicator']=='Y'].drop(columns=['Record Only Indicator', 'Claims above 50K', 'Total Incurred', 'Total Incurred + ALAE', 'Total Paid Loss Net Salvage/Subro/Loss Recovery', 'Incurred w/o ALAE', 'ALAE Reserve', 'ALAE Paid', 'Salvage Recovery', 'Subro Recovery', 'Loss Recovery', 'Deductible Recovery', 'Expense Recovery', 'Litigation Status', 'Outstanding Loss Reserve']).reset_index(drop=True)
    claim_df = df[df['Record Only Indicator']!='Y'].drop(columns=['Record Only Indicator']).reset_index(drop=True)

    
    if ro_df.empty:
        print("no record only data found")
        ro_df = pd.DataFrame(
            [["No Records Only"] + [None] * (len(ro_df.columns) - 1)],
            columns=ro_df.columns
        )


    # Setup file reader
    file_path = '/Volumes/claims_data_pre_prod/gold/statics/SACLossRunTemplate.xlsx'
    wb = load_workbook(file_path)

    # Write claim_df ----------------------------------------------------------------------
    print('writing claim data')

    # Get shape of data to write
    maxRow = claim_df.shape[0] + 1
    maxCol = claim_df.shape[1]
    maxRef = 'A1:'+get_column_letter(maxCol)+str(maxRow)

    # Select correct worksheet
    ws = wb['Claims Data']

    # Get table arrays for cell manipulation
    oldTblRange = ws.tables['ClaimsData'].ref
    oldTblRangeArray = ws[ws.tables['ClaimsData'].ref]
    newTblRangeArray = ws[maxRef]

    # Get info for trimming table should it shrink
    oldTblNumCols = len(oldTblRangeArray[0])
    oldTblNumRows = len(oldTblRangeArray)
    toDelRow = oldTblNumRows - maxRow if oldTblNumRows > maxRow else 0
    toDelCol = oldTblNumCols - maxCol if oldTblNumCols > maxCol else 0
    # print('toDel (row, col)=', toDelRow, toDelCol)
    # print('oldRange=',oldTblRange)
    # print('newRange=',maxRef)

    # Overwrite the data in the table
    print('writing new data (overwrite)')
    for rowRef,row in enumerate(newTblRangeArray):
        for colRef, cell in enumerate(row):
            toWrite = claim_df.columns[colRef] if rowRef == 0 else claim_df[claim_df.columns[colRef]][rowRef-1]
            # print('cell=,', cell, 'curVal=',cell.value,'valToWrite=',toWrite)
            cell.value = toWrite

    # Trim off any old data from overwrite (not necessary if new data was larger than old)
    if toDelRow != 0:
        print('deleting rows')
        for row in oldTblRangeArray[-toDelRow:]:
            # print('rowToDel=', row)
            for cell in row:
                # print('|\\__deleting cell', cell, 'setting value', cell.value, 'to ''None''')
                cell.value = None
                
    if toDelCol != 0:
        print('deleting cols')
        for row in oldTblRangeArray:
            # print('in row', row)
            for i in range(1,(toDelCol+1)):
                # print('|\\__deleting cell',row[-i],'cell value =',row[-i].value,'setting to ''None''')
                row[-i].value = None
                
    # change the table object columns
    columns = []
    for i, col in enumerate(claim_df.columns):
        columns.append(table.TableColumn(id=i+1, name=col))
    ws.tables['ClaimsData'].tableColumns = columns
    # Set table autofilter to None so as to prevent breaks
    ws.tables['ClaimsData'].autoFilter = None

    # update table object reference
    ws.tables['ClaimsData'].ref = maxRef

    # Write ro_df ----------------------------------------------------------------------
    print('writing record only data')

    # Get shape of data to write
    maxRow = ro_df.shape[0] + 1
    maxCol = ro_df.shape[1]
    maxRef = 'A1:'+get_column_letter(maxCol)+str(maxRow)

    # Select correct worksheet
    ws = wb['Record Only']

    # Get table arrays for cell manipulation
    oldTblRange = ws.tables['RecordOnlyData'].ref
    oldTblRangeArray = ws[ws.tables['RecordOnlyData'].ref]
    newTblRangeArray = ws[maxRef]

    # Get info for trimming table should it shrink
    oldTblNumCols = len(oldTblRangeArray[0])
    oldTblNumRows = len(oldTblRangeArray)
    toDelRow = oldTblNumRows - maxRow if oldTblNumRows > maxRow else 0
    toDelCol = oldTblNumCols - maxCol if oldTblNumCols > maxCol else 0
    # print('toDel (row, col)=', toDelRow, toDelCol)
    # print('oldRange=',oldTblRange)
    # print('newRange=',maxRef)

    # Overwrite the data in the table
    print('writing new data (overwrite)')
    for rowRef,row in enumerate(newTblRangeArray):
        for colRef, cell in enumerate(row):
            toWrite = ro_df.columns[colRef] if rowRef == 0 else ro_df[ro_df.columns[colRef]][rowRef-1]
            # print('cell=,', cell, 'curVal=',cell.value,'valToWrite=',toWrite)
            cell.value = toWrite

    # Trim off any old data from overwrite (not necessary if new data was larger than old)
    if toDelRow != 0:
        print('deleting rows')
        for row in oldTblRangeArray[-toDelRow:]:
            # print('rowToDel=', row)
            for cell in row:
                # print('|\\__deleting cell', cell, 'setting value', cell.value, 'to ''None''')
                cell.value = None
                
    if toDelCol != 0:
        print('deleting cols')
        for row in oldTblRangeArray:
            # print('in row', row)
            for i in range(1,(toDelCol+1)):
                # print('|\\__deleting cell',row[-i],'cell value =',row[-i].value,'setting to ''None''')
                row[-i].value = None
                
    # change the table object columns
    columns = []
    for i, col in enumerate(ro_df.columns):
        columns.append(table.TableColumn(id=i+1, name=col))
    ws.tables['RecordOnlyData'].tableColumns = columns
    # Set table autofilter to None so as to prevent breaks
    ws.tables['RecordOnlyData'].autoFilter = None

    # update table object reference
    ws.tables['RecordOnlyData'].ref = maxRef

    # update cover page -----------------------------------------------------------------
    print('updating cover page')

    todayStr = datetime.now().strftime("%m/%d/%Y")
    ws = wb['Cover Page']
    ws.cell(2,2, value = custNum)
    ws.cell(3,2, value = custName)
    ws.cell(4,2, value = todayStr)

    # Set Pivot to Refresh on Open
    # Assumes pivot is on its own sheet named 'Summary By Policy Year' and 'Chart'
    pivot_ws = [wb["Summary By Policy Year"] if "Summary By Policy Year" in wb.sheetnames else wb.worksheets[0], wb["Chart"] if "Chart" in wb.sheetnames else wb.worksheets[0]]

    for s in pivot_ws:
        for p in getattr(s, "_pivots", []):
            p.cache.refreshOnLoad = True

           
    wb.calculation.fullCalcOnLoad = True
    
    save_path = "/local_disk0/tmp/tempExcel.xlsx"
    wb.save(save_path)
    wb.close()
    dest_path_prefix = '/Volumes/claims_data_pre_prod/gold/external_volume/specialaccounts_lossruns_temporary/'
    dest = f"{dest_path_prefix}{custName}_{datetime.now().strftime('%Y_%m_%d')}.xlsx"
    print('SAVING LOSS RUN AT:',dest)
    copyfile(save_path,dest)
    print('FINISHED')

# COMMAND ----------

# DBTITLE 1,SQL Server Setup
specAcctsQuery = f"""
SELECT --TOP 10
a.CustomerNum,
a.CustomerName,
a.LossRunDistFreq--,
--d.RecipCat,
--d.DistVia,
--d.AttnTo,
--d.EMailAddress
FROM tblAcctSpecial a
--LEFT JOIN tblDistribute_LossRun d ON a.CustomerNum = d.CustomerNum
WHERE AcctStatus = 'Active' 
AND a.LossRunDistFreq <> 'Not Needed' 
AND a.LossRunDistFreq <> ''
--AND a.CustomerNum = '1506896261'
"""

lossRunQuery = f"""
SELECT * FROM SAC_Loss_Run
"""

jdbcHostname_azure = "clms-preprd-sqlmanagedinstance.3b98dc354c37.database.windows.net"
jdbcDatabase_azure= "CLMAA_SpecialAccounts"
jdbcPort_azure = 1433
SQLusername_azure = "ClaimsAzureSQLMIGA"
SQLPassword_azure = dbutils.secrets.get(scope="ClaimsAnalyticsDatabricksSecretScope", key="ClaimsAzureSQLMIGA")
jdbcUrl_azure = "jdbc:sqlserver://{0}:{1};database={2}".format(jdbcHostname_azure, jdbcPort_azure, jdbcDatabase_azure)
properties_azure = {
    "username" : SQLusername_azure,
    "password" : SQLPassword_azure,
    }


# COMMAND ----------

# DBTITLE 1,Filtered Data Setup
specialAccounts_sparkDF = spark.read \
  .format("jdbc") \
  .option("url", jdbcUrl_azure) \
  .option("query", specAcctsQuery) \
  .option("user", SQLusername_azure) \
  .option("password", SQLPassword_azure) \
  .load()

specialAccounts_sparkDF = specialAccounts_sparkDF.withColumn("CustomerName", regexp_replace(col("CustomerName"), "-", ""))

# after building your key list in Spark (small list)
keys = [(r[0:7]) for r in specialAccounts_sparkDF.select(["CustomerNum",
                                                          "CustomerName"
                                                          ]).distinct().collect()]
in_list = ",".join([f"'{k[0]}'" for k in keys])




# COMMAND ----------

# DBTITLE 1,Loss Run Query
output_sparkDF = spark.read \
  .format("jdbc") \
  .option("url", jdbcUrl_azure) \
  .option("query", lossRunQuery) \
  .option("user", SQLusername_azure) \
  .option("password", SQLPassword_azure) \
  .load()



# COMMAND ----------

from pyspark.sql.types import DecimalType, FloatType
for f in output_sparkDF.schema.fields:
    #print(f"Column {f.name} is {f.dataType}")
    if isinstance(f.dataType, DecimalType):
        #print(f"Converting column {f.name} to float")
        output_sparkDF = output_sparkDF.withColumn(f.name, col(f.name).cast(FloatType()))

output_sparkDF.dtypes
                                                    

# COMMAND ----------

output_df = output_sparkDF.toPandas()

# COMMAND ----------

# DBTITLE 1,Clean and Format DF
formatted_df = output_df.copy()
# Columns that should stay as strings (identifiers with leading zeros, codes, etc.)
keep_as_string = {"Customer Number", "Policy Insured Name", "Policy Number", "Policy MOD", "Claim Number", "Exposure", "Producing Agent Code", "Policy Holder Claim Number", "Vehicle Year"}
# Restore numeric types from any Java-to-string conversions, except string identifiers
for c in formatted_df.columns:
    if c not in keep_as_string:
        try:
            formatted_df[c] = pd.to_numeric(formatted_df[c])
        except ValueError:
            print(f"Couldn't convert column {c} to numeric")

# Format Exposure column as two digits, blank if null
if "Exposure" in formatted_df.columns:
    formatted_df["Exposure"] = formatted_df["Exposure"].apply(
        lambda x: f"{int(x):02d}" if pd.notnull(x) and str(x).strip() != "" else ""
    )

formatted_df['Distinct Claim Helper'] = (formatted_df.groupby('Claim Number').cumcount() == 0).astype(int)

print(f"Loaded {len(formatted_df)} rows with {len(formatted_df.columns)} columns")
formatted_df.head()

# COMMAND ----------

#formatted_df[formatted_df['Claim Number'] == '85-00898010']

# COMMAND ----------

# DBTITLE 1,Generate Loss Runs
# Process each customer with a single filter operation
for i, key in enumerate(keys):
    custNum, custName = key
    print(f"\nProcessing customer {i+1}/{len(keys)}: {custName}")
    
    # Filter the Pandas DataFrame for this specific customer
    dfPanda = formatted_df[output_df['Customer Number'] == custNum].reset_index(drop=True)
    
    # Add a way to calculate "distinct claims" in pivot table
    if len(dfPanda) > 0:
        print(f"  Found {len(dfPanda)} records")
        createLossRun(dfPanda, key)
    else:
        print(f"  No records found, skipping")